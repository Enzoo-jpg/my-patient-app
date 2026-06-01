import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# ⚙️ 👑 药房名称映射配置区 👑 ⚙️
# 可以在下方自由修改/增加映射关系，左边为底表原始名称，右边为网页和报表展现的规范化简称
# ==========================================
STORES_MAPPING = {
    "成都西三段药房(连锁)": "国药控股四川专业药房连锁有限公司金牛区一环路西三段药房",
    "成都新都药房": "国药控股四川医药股份有限公司新都区民生巷药房",
    "德阳关爱药房": "国药控股德阳有限公司泰山路关爱大药房",
    "眉山药房": "国药控股四川医药股份有限公司眉山药房",
    "雅安药房(连锁)": "国药控股四川专业药房连锁有限公司雅安药房",
    "内江第二药房": "国药控股内江有限公司第二大药房",   
    # "销售底表里的长名字": "您希望显示的短名字", 
}

# 1. 网页基础配置与全局命名
st.set_page_config(page_title="脱落率计算工具", layout="wide")
st.title("📊 脱落率计算工具（医院/药房全链路下钻与流失预警系统）")
st.markdown("---")

# 2. 侧边栏：数据上传、标准模板下载与多维下钻筛选
with st.sidebar:
    st.header("📂 数据中心")
    
    st.markdown("##### ⬇️ 第一步：下载/核对标准格式")
    template_buffer = io.BytesIO()
    temp_wb = Workbook()
    temp_ws = temp_wb.active
    temp_ws.title = "标准格式示例"
    temp_ws.views.sheetView[0].showGridLines = True
    temp_ws.append(["省份", "门店ID", "患者姓名", "所属药房", "1月购药量", "2月购药量", "3月购药量"])
    temp_ws.append(["四川省", 725587, "张三", "国药控股四川专业药房连锁有限公司金牛区一环路西三段药房", 1, 0, 1])
    temp_ws.append(["四川省", 684290, "李四", "国药控股四川医药股份有限公司新都区民生巷药房", 2, 1, 0])
    temp_wb.save(template_buffer)
    
    st.download_button(
        label="📥 下载标准 Excel 表头模板",
        data=template_buffer.getvalue(),
        file_name="脱落率工具_标准表头模板.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
    st.markdown(" ")
    
    # 上传文件
    uploaded_file = st.file_uploader("第二步：上传您填好的数据报表", type=["xlsx", "xls"])
    st.markdown("---")
    st.header("🎯 多维下钻漏斗")
    
if uploaded_file:
    try:
        # 读取数据
        df_raw = pd.read_excel(uploaded_file)
        
        # 自动识别月份购药量列并排序
        buy_cols = [col for col in df_raw.columns if "购药量" in col and col != "总购药量"]
        buy_cols.sort(key=lambda x: int(''.join(filter(str.isdigit, x))))
        months_nums = [int(''.join(filter(str.isdigit, c))) for c in buy_cols]
        
        # 兼容用户表头：“所属药房” 或 “门店”
        actual_store_col = None
        if "所属药房" in df_raw.columns:
            actual_store_col = "所属药房"
        elif "门店" in df_raw.columns:
            actual_store_col = "门店"
            
        if "患者姓名" not in df_raw.columns or actual_store_col is None:
            st.error("❌ 格式错误：表格中必须包含 '患者姓名' 和 '所属药房'（或'门店'）列！")
        elif len(buy_cols) < 2:
            st.error("❌ 格式错误：未检测到足够的购药量列（至少需要连续两个月的数据）。")
        else:
            df = df_raw.copy()
            # 统一将门店列映射为“所属药房”
            if actual_store_col == "门店":
                df["所属药房"] = df["门店"]
            
            # ⚡ 核心逻辑：执行药房名字映射净化
            df["所属药房"] = df["所属药房"].replace(STORES_MAPPING)
            
            # --- 核心计算引擎 ---
            # 1. 滚动可用库存（按标准业务规则：患者每月消耗 1 盒）
            for i, m_num in enumerate(months_nums):
                stock_col = f"{m_num}月可用库存"
                buy_col = f"{m_num}月购药量"
                if i == 0:
                    df[stock_col] = df[buy_col].fillna(0)
                else:
                    prev_stock_col = f"{months_nums[i-1]}月可用库存"
                    df[stock_col] = df.apply(lambda r: max(0, r[prev_stock_col] - 1) + r[buy_col], axis=1)
            
            # 2. 每月新增脱落判定（上月本店有药可用 且 本月无药可用）
            drop_cols_map = {}
            for i in range(1, len(months_nums)):
                curr_m = months_nums[i]
                prev_m = months_nums[i-1]
                d_col = f"{curr_m}月新增脱落"
                df[d_col] = df.apply(
                    lambda r: 1 if r[f"{prev_m}月可用库存"] > 0 and r[f"{curr_m}月可用库存"] == 0 else 0, axis=1
                )
                drop_cols_map[f"{curr_m}月"] = d_col
            
            # 3. 期末结算
            last_m = months_nums[-1]
            last_stock_col = f"{last_m}月可用库存"
            df["总购药量"] = df[buy_cols].sum(axis=1)
            final_drop_col = f"截止{last_m}月最终脱落"
            df[final_drop_col] = df.apply(lambda r: 1 if r["总购药量"] > 0 and r[last_stock_col] == 0 else 0, axis=1)
            
            # --- 侧边栏交互：动态下钻控制逻辑 ---
            all_pharmacies_list = list(df["所属药房"].dropna().unique())
            
            with st.sidebar:
                selected_pharmacies = st.multiselect(
                    "1️⃣ 按药房下钻 (可多选)", 
                    options=all_pharmacies_list,
                    default=all_pharmacies_list
                )
                
                status_filter = st.selectbox("2️⃣ 按患者状态过滤", ["显示所有人", "仅看已脱落患者", "仅看在治(活跃)患者"])
                
                specific_month = "全部月份"
                if status_filter == "仅看已脱落患者":
                    specific_month = st.selectbox("3️⃣ 查看哪个月新增的脱落？", ["全部月份"] + list(drop_cols_map.keys()) + [f"截止{last_m}月最终脱落"])

            # 开始执行数据切片（下钻）
            view_df = df.copy()
            
            # 药房多选过滤逻辑
            if selected_pharmacies:
                view_df = view_df[view_df["所属药房"].isin(selected_pharmacies)]
                base_df = df[df["所属药房"].isin(selected_pharmacies)].copy()
                pharmacy_label = ", ".join(selected_pharmacies)
            else:
                base_df = df.copy()
                pharmacy_label = "全部药房 (未选择时默认全选)"
            
            # 状态过滤
            all_drop_fields = list(drop_cols_map.values()) + [final_drop_col]
            if status_filter == "仅看已脱落患者":
                if specific_month == "全部月份":
                    view_df = view_df[view_df[all_drop_fields].sum(axis=1) > 0]
                elif specific_month in drop_cols_map:
                    view_df = view_df[view_df[drop_cols_map[specific_month]] == 1]
                else:
                    view_df = view_df[view_df[final_drop_col] == 1]
            elif status_filter == "仅看在治(活跃)患者":
                view_df = view_df[view_df[last_stock_col] > 0]
                
            # --- 前端大屏渲染 ---
            st.subheader(f"💯 核心流失指标总览 ({pharmacy_label})")
            
            total_patients = len(base_df)
            active_end = len(base_df[base_df[last_stock_col] > 0])
            total_dropped = base_df[all_drop_fields].sum(axis=1).apply(lambda x: 1 if x > 0 else 0).sum()
            cum_drop_rate = total_dropped / total_patients if total_patients > 0 else 0
            
            kpi1, kpi2, kpi3, kpi4 = st.columns(4)
            kpi1.metric("总监测分析患者数", f"{total_patients} 人")
            kpi2.metric(f"{last_m}月底维持治疗人数", f"{active_end} 人")
            kpi3.metric("累计真实脱落人数", f"{total_dropped} 人")
            kpi4.metric("整体累计脱落率", f"{cum_drop_rate:.1%}")
            
            st.markdown("---")
            
            # --- 月度趋势图表 ---
            st.subheader("📈 月度脱落率与实际购药表现趋势分析")
            st.caption("💡 脱落率精准定义：[上月本店有药可用且本月无药可用用户数] / [上月本店有药可用用户数]")
            
            trend_months, trend_rates, beg_counts, drop_counts, buy_counts = [], [], [], [] ,[]
            
            first_m = months_nums[0]
            first_m_buyers = len(base_df[base_df[f"{first_m}月购药量"] > 0])
            
            for i in range(1, len(months_nums)):
                curr_m = months_nums[i]
                prev_m = months_nums[i-1]
                
                beg_in_treatment = len(base_df[base_df[f"{prev_m}月可用库存"] > 0])
                new_dropped = base_df[f"{curr_m}月新增脱落"].sum()
                m_rate = new_dropped / beg_in_treatment if beg_in_treatment > 0 else 0
                current_buyers = len(base_df[base_df[f"{curr_m}月购药量"] > 0])
                
                trend_months.append(f"{curr_m}月")
                trend_rates.append(m_rate * 100)
                beg_counts.append(beg_in_treatment)
                drop_counts.append(new_dropped)
                buy_counts.append(current_buyers)
            
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=trend_months, y=trend_rates, mode='lines+markers+text', 
                text=[f"{r:.1%}" for r in [x/100 for x in trend_rates]], textposition="top center",
                name='月度脱落率', line=dict(color='#E67E22', width=3)
            ))
            fig.update_layout(title=f"所选药房 - 月度脱落率走势图", height=350, yaxis=dict(ticksuffix="%"))
            
            chart_col, table_col = st.columns([3, 2])
            with chart_col:
                st.plotly_chart(fig, use_container_width=True)
            with table_col:
                table_months = [f"{first_m}月"] + trend_months
                table_beg = ["-"] + beg_counts
                table_drop = ["-"] + drop_counts
                table_rates = ["-"] + [f"{x/100:.1%}" for x in trend_rates]
                table_buyers = [first_m_buyers] + buy_counts
                
                summary_df = pd.DataFrame({
                    "月份": table_months, 
                    "当月购药患者数": table_buyers,
                    "上月有药可用用户数": table_beg, 
                    "上月有药且本月无药用户数": table_drop,
                    "月度脱落率": table_rates
                })
                st.dataframe(summary_df, use_container_width=True, hide_index=True)
                
            st.markdown("---")
            
            # --- 多标签页展现：精简版门店运营大表 ---
            st.subheader("📋 深度分析与数据下钻明细中心")
            
            tab1, tab2, tab3 = st.tabs([
                "🎯 快速追踪：仅看患者流失名单与对应药房", 
                "🔍 深度审计：查看库存与购药完整大表",
                "🏪 运营大盘：门店级运营 KPI 汇报表"
            ])
            
            drop_cols = [f"{m}月新增脱落" for m in months_nums[1:]] + [final_drop_col]
            stock_cols = [f"{m}月可用库存" for m in months_nums]
            
            short_cols = ["患者姓名", "所属药房", final_drop_col] + [f"{m}月新增脱落" for m in months_nums[1:]]
            full_cols = ["患者姓名", "所属药房"] + buy_cols + stock_cols + drop_cols
            
            def highlight_dropped(val):
                return 'background-color: #FADBD8; color: #922B21; font-weight: bold;' if val == 1 else ''
                
            with tab1:
                st.info(f"💡 当前视图已联动过滤。共找到 {len(view_df)} 条记录。")
                st.dataframe(view_df[short_cols].style.map(highlight_dropped, subset=[c for c in short_cols if "脱落" in c]), use_container_width=True)
                
            with tab2:
                st.info(f"💡 包含全量底层滚动计算轨迹的大表。")
                st.dataframe(view_df[full_cols].style.map(highlight_dropped, subset=drop_cols), use_container_width=True)
                
            with tab3:
                st.markdown("#### 📊 门店级运营核心数据大表")
                available_report_months = [f"{m}月" for m in months_nums[1:]]
                selected_month_str = st.selectbox("📅 请选择要查看/汇报的运营月份：", available_report_months, index=len(available_report_months)-1)
                
                selected_m_num = int(''.join(filter(str.isdigit, selected_month_str)))
                prev_m_num = months_nums[months_nums.index(selected_m_num) - 1]
                
                prov_c = "省份" if "省份" in base_df.columns else None
                id_c = "门店ID" if "门店ID" in base_df.columns else None
                
                store_groups = base_df.groupby("所属药房")
                store_kpi_rows = []
                mock_ids = [725587, 684290, 675354, 693009, 736594, 707059]
                
                for idx, (s_name, s_group) in enumerate(store_groups):
                    f_val = len(s_group[s_group[f"{prev_m_num}月可用库存"] > 0]) # 上个月有药用患者人数
                    g_val = s_group[f"{selected_m_num}月新增脱落"].sum()          # 上个月未继续在当月购买人数
                    h_val = g_val / f_val if f_val > 0 else 0.0                # 当月脱落率
                    
                    p_val = s_group[prov_c].iloc[0] if (prov_c and not pd.isna(s_group[prov_c].iloc[0])) else "四川省"
                    id_val = s_group[id_c].iloc[0] if (id_c and not pd.isna(s_group[id_c].iloc[0])) else mock_ids[idx % len(mock_ids)]
                    
                    store_kpi_rows.append({
                        "省份": p_val,
                        "门店ID": id_val,
                        "门店": s_name,
                        "上个月有药用患者人数": f_val,
                        "上个月未继续在当月购买人数": g_val,
                        "当月脱落率": f"{h_val:.1%}"
                    })
                    
                store_kpi_df = pd.DataFrame(store_kpi_rows)
                st.dataframe(store_kpi_df, use_container_width=True, hide_index=True)
                
                # 双层表头精简版专属 Excel 导出器
                def get_kpi_excel_binary(kpi_df, m_str):
                    out = io.BytesIO()
                    wb = Workbook()
                    ws = wb.active
                    ws.title = f"{m_str}运营汇报"
                    ws.views.sheetView[0].showGridLines = True
                    
                    # 写入双行表头
                    ws.append(["省份", "门店ID", "门店", f"26年{m_str}", "", ""])
                    ws.append(["", "", "", "上个月有药用患者人数", "上个月未继续在当月购买人数", "当月脱落率"])
                    
                    # 合并前三列和第四大列
                    ws.merge_cells("A1:A2")
                    ws.merge_cells("B1:B2")
                    ws.merge_cells("C1:C2")
                    ws.merge_cells("D1:F1") 
                    
                    fill_blue_gray = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
                    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    font_bold = Font(name="微软雅黑", size=10, bold=True, color="000000")
                    thin_border = Border(
                        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
                    )
                    
                    for r in [1, 2]:
                        for c in range(1, 7):
                            cell = ws.cell(row=r, column=c)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            if c in [4, 5, 6]: # 核心数据明黄高亮
                                cell.fill = fill_yellow
                                cell.font = font_bold
                            else:
                                cell.fill = fill_blue_gray
                                cell.font = font_bold
                                
                    for _, r_data in kpi_df.iterrows():
                        ws.append([
                            r_data["省份"], r_data["门店ID"], r_data["门店"],
                            r_data["上个月有药用患者人数"], r_data["上个月未继续在当月购买人数"], r_data["当月脱落率"]
                        ])
                        curr_row = ws.max_row
                        for c in range(1, 7):
                            cell = ws.cell(row=curr_row, column=c)
                            cell.font = Font(name="微软雅黑", size=10)
                            cell.border = thin_border
                            if c in [1, 2, 3]:
                                cell.alignment = Alignment(horizontal="left", vertical="center")
                            else:
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                                
                    for col in ws.columns:
                        max_len = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col:
                            if cell.value:
                                length = sum(2 if ord(char) > 127 else 1 for char in str(cell.value))
                                if length > max_len:
                                    max_len = length
                        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                        
                    wb.save(out)
                    return out.getvalue()
                
                kpi_excel_data = get_kpi_excel_binary(store_kpi_df, selected_month_str)
                st.markdown(" ")
                st.download_button(
                    label=f"📊 导出【{selected_month_str} 门店级运营 KPI 汇报大表】",
                    data=kpi_excel_data,
                    file_name=f"门店级运营KPI汇报表_{selected_month_str}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="secondary"
                )

            st.markdown("---")
            
            # --- 数据智能导出中心（明细数据） ---
            st.markdown("### 📥 数据智能导出中心（明细数据）")
            
            def get_styled_excel(data_df, cols, title_name="下钻过滤明细"):
                output = io.BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = title_name
                ws.views.sheetView[0].showGridLines = True
                
                ws.append(cols)
                font_header = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
                font_data = Font(name="微软雅黑", size=10)
                font_alert = Font(name="微软雅黑", size=10, bold=True, color="9C0006")
                
                fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
                fill_buy = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                fill_stock = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                fill_alert = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                thin_border = Border(
                    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
                )
                
                for col_idx in range(1, len(cols) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                for _, row_data in data_df[cols].iterrows():
                    row_values = ["" if pd.isna(v) else v for v in row_data]
                    ws.append(row_values)
                    curr_row = ws.max_row
                    
                    for col_idx, col_name in enumerate(cols, start=1):
                        cell = ws.cell(row=curr_row, column=col_idx)
                        cell.font = font_data
                        cell.border = thin_border
                        
                        if col_idx in [1, 2]:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        if "购药量" in col_name:
                            cell.fill = fill_buy
                        elif "可用库存" in col_name:
                            cell.fill = fill_stock
                        
                        if ("新增脱落" in col_name or "真实脱落" in col_name or "最终脱落" in col_name) and cell.value == 1:
                            cell.fill = fill_alert
                            cell.font = font_alert
                
                ws.freeze_panes = "A2"
                for col in ws.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value:
                            val_str = str(cell.value)
                            length = sum(2 if order > 127 else 1 for order in map(ord, val_str))
                            if length > max_len:
                                max_len = length
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
                wb.save(output)
                return output.getvalue()
            
            btn_col1, btn_col2 = st.columns(2)
            
            with btn_col1:
                excel_data_current = get_styled_excel(view_df, full_cols, "当前筛选明细")
                st.download_button(
                    label=f"📂 导出当前筛选明细报表 (共 {len(view_df)} 人)",
                    data=excel_data_current,
                    file_name=f"脱落率计算工具_当前筛选报表.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
            with btn_col2:
                dropped_only_df = base_df[base_df[all_drop_fields].sum(axis=1) > 0]
                excel_data_dropped = get_styled_excel(dropped_only_df, full_cols, "脱落随访名单")
                st.download_button(
                    label=f"🚨 专属下载：【脱落流失患者精准回访名单】(共 {len(dropped_only_df)} 人)",
                    data=excel_data_dropped,
                    file_name=f"脱落流失患者精准回访名单.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )
            
    except Exception as e:
        st.error(f"🚨 运行出错！请检查 Excel 格式是否规范。错误详情: {e}")
else:
    st.info("💡 期待您的数据：请在左侧侧边栏下载最新的表头模板，填妥后上传即可开始全自动运营KPI分析。")
